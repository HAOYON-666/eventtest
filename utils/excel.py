import os, shutil, json, random, sys, logging, glob
from pprint import pprint
import pandas as pd
import numpy as np
import re
import copy
import openpyxl
from collections import defaultdict
from easydict import EasyDict as edict
from .infer_result import GatherRes

class WriteExcel():
    
    def __init__(self, file_path):
        self.file_path = file_path
        if not os.path.exists(file_path):
            self._init_excel()

    def _init_excel(self):
        # 初始化excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "总表"
        wb.create_sheet(title="子表统计")
        wb.create_sheet(title="大模型修正")
        wb.save(self.file_path)

    def fresh(self, sheet_name, data):
        # 子表追加
        ori = pd.read_excel(self.file_path, sheet_name=None)
        ori_sheet_names = ori.keys()    
        if sheet_name in ori_sheet_names: # 有子表拼接原先的数据
            data = pd.DataFrame(data)
            new_sheet = pd.concat([ori[sheet_name], data], ignore_index=True)
        else:
            new_sheet = pd.DataFrame(data)
        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as write:
            new_sheet.to_excel(write, sheet_name=sheet_name, index=False)
        self._fresh_main()  
        self._fresh_llm()

    def _fresh_llm(self):
        main_sheet = {
            "事件类型" : [],
            "事件ID" : [],
            "场景" : [],
            "成熟度" : [],
            "视频数量": [],
            "正报视频数量" : [],
            "正报视频比例" : [],
            "正报数量" : [],
            "正报重复" : [],
            "误报数量" : [],
            "误报重复" : [],
            "事件数量" : [],
            "事件正确率" : [],
            "是否通过" : [],
        }
        NiXing = edict({
            "视频数量": 0,
            "正报视频数量" : 0,
            "正报数量" : 0,
            "正报重复" : 0,
            "误报数量" : 0,
            "误报重复" : 0,
        })
        NiXing_flag = False
        sub_sheets = self._static_sub_sheet()
        for event_type, sub_sheet in sub_sheets.items():
            for row in sub_sheet.iterrows():
                sub_res  = self.parse_sub_sheet(row)
                # 合并逆行的结果
                if sub_res.event_id in ["256", "257", "258", "259"]:
                    NiXing["视频数量"] += sub_res.video_num
                    NiXing["正报视频数量"] += sub_res.llm_right_video_num
                    NiXing["正报数量"] += sub_res.llm_right_event_num
                    NiXing["正报重复"] += sub_res.llm_right_event_repeat_num
                    NiXing["误报数量"] += sub_res.llm_wrong_event_num
                    NiXing["误报重复"] += sub_res.llm_wrong_event_repeat_num
                    NiXing_flag = True
                else:
                    main_sheet["事件类型"].append(event_type)
                    main_sheet["事件ID"].append(sub_res.event_id)
                    main_sheet["场景"].append(sub_res.scene)
                    main_sheet["成熟度"].append(None)
                    main_sheet["视频数量"].append(sub_res.video_num)
                    main_sheet["正报视频数量"].append(sub_res.llm_right_video_num)
                    main_sheet["正报视频比例"].append(self._get_percent(sub_res.llm_right_video_num, sub_res.video_num))
                    main_sheet["正报数量"].append(sub_res.llm_right_event_num)
                    main_sheet["正报重复"].append(sub_res.llm_right_event_repeat_num)
                    main_sheet["误报数量"].append(sub_res.llm_wrong_event_num)
                    main_sheet["误报重复"].append(sub_res.llm_wrong_event_repeat_num)
                    main_sheet["事件数量"].append(sub_res.llm_event_num)
                    main_sheet["事件正确率"].append(self._get_percent(sub_res.llm_right_event_num + sub_res.llm_right_event_repeat_num, sub_res.llm_event_num))
                    main_sheet["是否通过"].append(sub_res.llm_right_or_not)
        # 逆行的结果
        if NiXing_flag:
            main_sheet["事件类型"].append("逆行")
            main_sheet["事件ID"].append("256-259")
            main_sheet["场景"].append(None)
            main_sheet["成熟度"].append(None)
            main_sheet["视频数量"].append(NiXing["视频数量"])
            main_sheet["正报视频数量"].append(NiXing["正报视频数量"])
            main_sheet["正报视频比例"].append(self._get_percent(NiXing["正报视频数量"], NiXing["视频数量"]))
            main_sheet["正报数量"].append(NiXing["正报数量"])
            main_sheet["正报重复"].append(NiXing["正报重复"])
            main_sheet["误报数量"].append(NiXing["误报数量"])
            main_sheet["误报重复"].append(NiXing["误报重复"])
            all_event_num = NiXing["正报数量"] + NiXing["正报重复"] + NiXing["误报数量"] + NiXing["误报重复"]
            main_sheet["事件数量"].append(all_event_num)
            main_sheet["事件正确率"].append(self._get_percent(NiXing["正报数量"] + NiXing["正报重复"], all_event_num))
            main_sheet["是否通过"].append(1 if NiXing["正报视频数量"]/(NiXing["视频数量"] + 1e-5) >= 0.8 else 0)

        main_sheet = self._add_empty_row(main_sheet)
        main_sheet["事件类型"].append("通过率")
        main_sheet["事件ID"].append(f'{len([i for i in main_sheet["是否通过"] if i == 1])}/{len(main_sheet["是否通过"])-1}')
        main_sheet["场景"].append(f'{round(len([i for i in main_sheet["是否通过"] if i == 1])/(len(main_sheet["是否通过"])-1 + 1e-5)*100, 2)}%')
        main_sheet["成熟度"].append(None)
        main_sheet["视频数量"].append(None)
        main_sheet["正报视频数量"].append(None)
        main_sheet["正报视频比例"].append(None)
        main_sheet["正报数量"].append(None)
        main_sheet["正报重复"].append(None)
        main_sheet["误报数量"].append(None)
        main_sheet["误报重复"].append(None)
        main_sheet["事件数量"].append(None)
        main_sheet["事件正确率"].append(None)
        main_sheet["是否通过"].append(None)

        gather_all = pd.DataFrame(main_sheet)
        gather_all = self._add_bg_color(gather_all)
        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as write:
            gather_all.to_excel(write, sheet_name="大模型修正", index=False)


    def _fresh_main(self):
        '''更新子表统计 
        Args:
            arg
        Outputs:
            out
        '''
        main_sheet = {
            "事件类型" : [],
            "事件ID" : [],
            "场景" : [],
            "成熟度" : [],
            "视频数量": [],
            "正报视频数量" : [],
            "正报视频比例" : [],
            "正报数量" : [],
            "正报重复" : [],
            "误报数量" : [],
            "误报重复" : [],
            "事件数量" : [],
            "事件正确率" : [],
            "是否通过" : [],
        }
        NiXing = edict({
            "视频数量": 0,
            "正报视频数量" : 0,
            "正报数量" : 0,
            "正报重复" : 0,
            "误报数量" : 0,
            "误报重复" : 0,
        })
        NiXing_flag = False
        sub_sheets = self._static_sub_sheet()
        for event_type, sub_sheet in sub_sheets.items():
            for row in sub_sheet.iterrows():
                sub_res  = self.parse_sub_sheet(row)
                # 合并逆行的结果
                if sub_res.event_id in ["256", "257", "258", "259"]:
                    NiXing["视频数量"] += sub_res.video_num
                    NiXing["正报视频数量"] += sub_res.right_video_num
                    NiXing["正报数量"] += sub_res.right_event_num
                    NiXing["正报重复"] += sub_res.right_event_repeat_num
                    NiXing["误报数量"] += sub_res.wrong_event_num
                    NiXing["误报重复"] += sub_res.wrong_event_repeat_num
                    NiXing_flag = True
                else:
                    main_sheet["事件类型"].append(event_type)
                    main_sheet["事件ID"].append(sub_res.event_id)
                    main_sheet["场景"].append(sub_res.scene)
                    main_sheet["成熟度"].append(None)
                    main_sheet["视频数量"].append(sub_res.video_num)
                    main_sheet["正报视频数量"].append(sub_res.right_video_num)
                    main_sheet["正报视频比例"].append(self._get_percent(sub_res.right_video_num, sub_res.video_num))
                    main_sheet["正报数量"].append(sub_res.right_event_num)
                    main_sheet["正报重复"].append(sub_res.right_event_repeat_num)
                    main_sheet["误报数量"].append(sub_res.wrong_event_num)
                    main_sheet["误报重复"].append(sub_res.wrong_event_repeat_num)
                    main_sheet["事件数量"].append(sub_res.event_num)
                    main_sheet["事件正确率"].append(self._get_percent(sub_res.right_event_num + sub_res.right_event_repeat_num, sub_res.event_num))
                    main_sheet["是否通过"].append(sub_res.right_or_not)
        # 逆行的结果
        if NiXing_flag:
            main_sheet["事件类型"].append("逆行")
            main_sheet["事件ID"].append("256-259")
            main_sheet["场景"].append(None)
            main_sheet["成熟度"].append(None)
            main_sheet["视频数量"].append(NiXing["视频数量"])
            main_sheet["正报视频数量"].append(NiXing["正报视频数量"])
            main_sheet["正报视频比例"].append(self._get_percent(NiXing["正报视频数量"], NiXing["视频数量"]))
            main_sheet["正报数量"].append(NiXing["正报数量"])
            main_sheet["正报重复"].append(NiXing["正报重复"])
            main_sheet["误报数量"].append(NiXing["误报数量"])
            main_sheet["误报重复"].append(NiXing["误报重复"])
            all_event_num = NiXing["正报数量"] + NiXing["正报重复"] + NiXing["误报数量"] + NiXing["误报重复"]
            main_sheet["事件数量"].append(all_event_num)
            main_sheet["事件正确率"].append(self._get_percent(NiXing["正报数量"] + NiXing["正报重复"], all_event_num))
            main_sheet["是否通过"].append(1 if NiXing["正报视频数量"]/(NiXing["视频数量"] + 1e-5) >= 0.8 else 0)

        # 复制结果用于统计总表
        main_sheet_copy = copy.deepcopy(main_sheet)

        main_sheet = self._add_empty_row(main_sheet)
        main_sheet["事件类型"].append("通过率")
        main_sheet["事件ID"].append(f'{len([i for i in main_sheet["是否通过"] if i == 1])}/{len(main_sheet["是否通过"])-1}')
        main_sheet["场景"].append(f'{round(len([i for i in main_sheet["是否通过"] if i == 1])/(len(main_sheet["是否通过"])-1 + 1e-5)*100, 2)}%')
        main_sheet["成熟度"].append(None)
        main_sheet["视频数量"].append(None)
        main_sheet["正报视频数量"].append(None)
        main_sheet["正报视频比例"].append(None)
        main_sheet["正报数量"].append(None)
        main_sheet["正报重复"].append(None)
        main_sheet["误报数量"].append(None)
        main_sheet["误报重复"].append(None)
        main_sheet["事件数量"].append(None)
        main_sheet["事件正确率"].append(None)
        main_sheet["是否通过"].append(None)

        gather_all = pd.DataFrame(main_sheet)
        gather_all = self._add_bg_color(gather_all)
        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as write:
            gather_all.to_excel(write, sheet_name="子表统计", index=False)

        # 更新总表
        new_sheet = {
            "事件类型" : [],
            "事件ID" : [],
            "场景" : [],
            "成熟度" : [],
            "视频数量": [],
            "正报视频数量" : [],
            "正报视频比例" : [],
            "正报数量" : [],
            "正报重复" : [],
            "误报数量" : [],
            "误报重复" : [],
            "事件数量" : [],
            "事件正确率" : [],
            "是否通过" : [],
        }
        for event in set(main_sheet_copy["事件类型"]):
            idx = [i == event for i in main_sheet_copy["事件类型"]]
            sub = {k : [v[i] for i, j in enumerate(idx) if j] for k, v in main_sheet_copy.items()}

            new_sheet["事件类型"].append(event)
            new_sheet["事件ID"].append(sub["事件ID"][0])
            new_sheet["场景"].append(None)
            new_sheet["成熟度"].append(None)
            new_sheet["视频数量"].append(sum(sub["视频数量"]))
            new_sheet["正报视频数量"].append(sum(sub["正报视频数量"]))
            new_sheet["正报视频比例"].append(self._get_percent(sum(sub["正报视频数量"]), sum(sub["视频数量"])))
            new_sheet["正报数量"].append(sum(sub["正报数量"]))
            new_sheet["正报重复"].append(sum(sub["正报重复"]))
            new_sheet["误报数量"].append(sum(sub["误报数量"]))
            new_sheet["误报重复"].append(sum(sub["误报重复"]))
            new_sheet["事件数量"].append(sum(sub["事件数量"]))
            new_sheet["事件正确率"].append(self._get_percent(sum(sub["正报数量"]) + sum(sub["正报重复"]), sum(sub["事件数量"])))
            new_sheet["是否通过"].append(1 if (sum(sub["正报数量"]) + sum(sub["正报重复"]))/(sum(sub["事件数量"])+1e-5) > 0.8 else 0)  
        new_sheet = self._add_empty_row(new_sheet)
        gather_all_new = pd.DataFrame(new_sheet)
        gather_all_new = self._add_bg_color(gather_all_new)
        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as write:
            gather_all_new.to_excel(write, sheet_name="总表", index=False)

    def _add_bg_color(self, gather_all):
        '''为总表的数值添加背景颜色 
        Args:
            arg
        Outputs:
            out
        '''
        # 生成颜色渐变
        start_color = [255, 0, 0]  
        end_color = [0, 255, 0]    
        colors = []
        num_colors = 102
        for i in range(num_colors):
            r = int(start_color[0] + (end_color[0] - start_color[0]) * (i / (num_colors - 1)))
            g = int(start_color[1] + (end_color[1] - start_color[1]) * (i / (num_colors - 1)))
            b = int(start_color[2] + (end_color[2] - start_color[2]) * (i / (num_colors - 1)))
            color_hex = "#{:02X}{:02X}{:02X}".format(r, g, b)
            colors.append(color_hex)

        def color(val):
            result = val == 1
            ans = ["background-color: green" if v else "background-color: red" for v in result]
            return ans[:-2] + ["background-color: None"]*2
        
        def percent_color(val):
            if "%" in str(val.iloc[0]):
                value = [int(float(i.split("%")[0])) for i in val[:-2]]
                ans = [f"background-color: {colors[i]}" for i in value]
            else:
                result = val[:-2] == 1
                ans = [f"background-color: {colors[-1]}" if v else f"background-color: {colors[0]}" for v in result]
            return ans + ["background-color: None"]*2
        
        # gather_all = gather_all.apply(color, subset=['是否通过'])
        gather_all = gather_all.style.apply(percent_color, subset=['正报视频比例', '事件正确率', "是否通过"])
        return gather_all

    def _get_percent(self, num, total): 
        '''获取百分比 
        Args:
            arg
        Outputs:
            out
        '''
        return f'{round(float(num)/(float(total)+1e-5)*100, 2)}%'


    def _add_empty_row(self, main_sheet):
        '''添加空行
        Args:
            arg
        Outputs:
            out
        '''
        for k, v in main_sheet.items():
            v.append(None)
        return main_sheet

    def parse_sub_sheet(self, sheet:pd.DataFrame, score_threshold=0.8):
        '''解析子表格
        Args:
            sheet: pd.DataFrame:子表格中的一行
        Outputs:
            out
        '''
        sub_res = edict()
        sheet = sheet[1]
        sub_res.event_id = eval(sheet["参数"])["EventType"]
        # event_id = re.search(r'EventType\s*:\s*(\d+)', event_id).group(1)
        sub_res.scene = sheet["场景"]  
        
        sub_static = eval(sheet["视频名"])   
        right_rate = sub_static["正确报警视频"]
        sub_res.video_num = int(right_rate.split("/")[1])
        sub_res.right_video_num = int(right_rate.split("/")[0])

        sub_res.right_event_num = sub_static["正确"]
        sub_res.right_event_repeat_num = sub_static["正确重复"]
        sub_res.wrong_event_num = sub_static["错误"]
        sub_res.wrong_event_repeat_num = sub_static["错误重复"]
        
        sub_res.event_num = sub_res.right_event_num + sub_res.wrong_event_num + sub_res.right_event_repeat_num + sub_res.wrong_event_repeat_num

        sub_res.llm_right_video_num = int(sub_static["llm正确报警视频"].split("/")[0])
        sub_res.llm_right_event_num = sub_static["llm正确"]
        sub_res.llm_right_event_repeat_num = sub_static["llm正确重复"]
        sub_res.llm_wrong_event_num = sub_static["llm错误"]
        sub_res.llm_wrong_event_repeat_num = sub_static["llm错误重复"]
        sub_res.llm_event_num = sub_res.llm_right_event_num + sub_res.llm_wrong_event_num + sub_res.llm_right_event_repeat_num + sub_res.llm_wrong_event_repeat_num

        
        if sub_res.right_video_num/sub_res.video_num >= score_threshold:
            sub_res.right_or_not = 1
        else:
            sub_res.right_or_not = 0

        if sub_res.llm_right_video_num/sub_res.video_num >= score_threshold:
            sub_res.llm_right_or_not = 1
        else:
            sub_res.llm_right_or_not = 0

        return sub_res

    def _static_sub_sheet(self):
        '''遍历所有子表格的统计结果 
        Args:
            arg
        Outputs:
            res: defaultdic : {event_type : [此事件类型下的所有子表格的‘=’行,...]}
        '''
        res = {}
        ori = pd.read_excel(self.file_path, sheet_name=None)
        for event, frame in ori.items():
            if event == "总表" or event == "Sheet1" or event == "子表统计" or event == "大模型修正":
                continue
            res[event] =  frame[frame["报警数量"] == "="]
        return res
    


if __name__=='__main__':
    import yaml
    from configInfo import ConfigInfo

    def read_conf(conf_path):
        with open(conf_path, "r") as f:
            g_config = edict(yaml.safe_load(f.read()))
        return g_config
    g_config = read_conf("../gather_lh_96.yaml")
    c = ConfigInfo("/data/liuhui_test/test_data/牛皮纸箱/9/configInfo.json", g_config)
    test_config = c.get_sheet_config()
    log_paths = glob.glob("/data/liuhui_test/result/20240605-110533_result_牛皮纸箱_9/*event_llm.log")
    test_dir = "/data/liuhui_test/result/20240605-110533_result_牛皮纸箱_9"
    gather_res = GatherRes(log_paths=log_paths, test_dir=test_dir, test_config=test_config, g_config=g_config)
    import ipdb; ipdb.set_trace()
    df = gather_res.get_df()
    gather_res.move_video()
    e = WriteExcel("./tmp.xlsx")
    e.fresh(sheet_name="test", data=df)
